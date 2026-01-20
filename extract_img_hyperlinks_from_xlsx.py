from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

XLSX_PATH = "Image_Captioning_For_Conflict_Awareness (1).xlsx"   # change if needed
SHEET_NAME = None  # None = active sheet
OUT_CSV = "dataset_with_links.csv"

# Column names in your sheet
COL_ID = "ID"
COL_IMG_NAME = "Img Name"
COL_IMG_URL = "Img URL"  # display text, but hyperlink target may exist

wb = load_workbook(XLSX_PATH)
ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

# Read header row -> column index
headers = [cell.value for cell in ws[1]]
col_map = {h: i+1 for i, h in enumerate(headers)}

missing_cols = [c for c in [COL_ID, COL_IMG_NAME, COL_IMG_URL] if c not in col_map]
if missing_cols:
    raise ValueError(f"Missing columns in xlsx: {missing_cols}. Found: {headers}")

rows = []
for r in range(2, ws.max_row + 1):
    id_val = ws.cell(r, col_map[COL_ID]).value
    img_name = ws.cell(r, col_map[COL_IMG_NAME]).value
    cell = ws.cell(r, col_map[COL_IMG_URL])

    display = cell.value
    link = None
    if cell.hyperlink is not None:
        # target is the actual clickable URL
        link = cell.hyperlink.target

    rows.append({
        "ID": id_val,
        "Img Name": img_name,
        "Img URL (display)": display,
        "Img URL (hyperlink)": link
    })

df = pd.DataFrame(rows)
df.to_csv(OUT_CSV, index=False)

print("Saved:", OUT_CSV)
print("Hyperlinks found:", df["Img URL (hyperlink)"].notna().sum(), "/", len(df))
print("Example hyperlinks:", df["Img URL (hyperlink)"].dropna().head(5).tolist())
